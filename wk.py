# coding=gbk
import os
from openpyxl import load_workbook, Workbook

if os.path.exists('result.xlsx'):
    os.remove('result.xlsx')

rb = Workbook()
rb.save('result.xlsx')
rs = rb["Sheet"]

wb = load_workbook("workdata.xlsx")

ws = wb["Sheet1"]
i = 1
# 自定义更改查询条件
name = 'C12H26'
col = 0  # 记目标
data = ['Timestep', name]
rs.append(data)
nums = [0, 0]
is_target = False
while i < 10001:
    for cell in ws[i]:
        if cell.value == name:
            col = cell.column
            is_target = True
    i += 1
    if is_target:
        for num in ws[i]:
            if num.column == 1:
                nums[0] = num.value
            if num.column == col:
                nums[1] = num.value
        rs.append(nums)
        rb.save('result.xlsx')
        is_target = False
